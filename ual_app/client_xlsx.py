import csv
import io
import re
from pathlib import Path
from typing import List, Tuple


MAX_EXCEL_ROWS = 1_048_576
MAX_EXCEL_COLUMNS = 16_384
ILLEGAL_XML = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _clean(value) -> str:
    return ILLEGAL_XML.sub("", str(value or ""))


def _unique_headers(values: List[str]) -> List[str]:
    headers, used = [], set()
    for index, value in enumerate(values, 1):
        base = _clean(value).strip() or f"Column {index}"
        candidate, suffix = base, 2
        while candidate.casefold() in used:
            candidate = f"{base} ({suffix})"
            suffix += 1
        used.add(candidate.casefold())
        headers.append(candidate)
    return headers


def clean_csv(raw: bytes) -> Tuple[List[str], List[List[str]]]:
    text = raw.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel
    source = [[_clean(value) for value in row] for row in csv.reader(io.StringIO(text), dialect=dialect)]
    source = [row for row in source if any(value.strip() for value in row)]
    if len(source) < 2:
        raise ValueError("The CSV must contain a header and at least one data row")
    width = max(len(row) for row in source)
    if width > MAX_EXCEL_COLUMNS:
        raise ValueError("The CSV exceeds Excel's 16,384-column limit")
    source = [row + [""] * (width - len(row)) for row in source]
    data_rows = source[1:]
    populated_columns = [index for index in range(width) if any(row[index].strip() for row in data_rows)]
    if not populated_columns:
        raise ValueError("The CSV contains no data values")
    headers = _unique_headers([source[0][index] for index in populated_columns])
    rows = [[row[index] for index in populated_columns] for row in data_rows]
    rows = [row for row in rows if any(value.strip() for value in row)]
    if len(rows) + 1 > MAX_EXCEL_ROWS:
        raise ValueError("The CSV exceeds Excel's 1,048,576-row limit")
    return headers, rows


def client_xlsx_bytes(filename: str, raw: bytes) -> Tuple[bytes, str]:
    if Path(filename).suffix.lower() not in (".csv", ".txt"):
        raise ValueError("Choose a CSV file")
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise ValueError("Client XLSX conversion requires openpyxl; run: pip install -r requirements.txt") from exc

    headers, rows = clean_csv(raw)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Client Data"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    sheet.freeze_panes = "A2"
    sheet.sheet_format.zeroHeight = True

    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.data_type = "s"
    for row_number, values in enumerate(rows, 2):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_number, column, value)
            cell.data_type = "s"

    last_row, last_column = len(rows) + 1, len(headers)
    last_letter = get_column_letter(last_column)
    table = Table(displayName="ClientShareableData", ref=f"A1:{last_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.row_dimensions[1].height = 24
    for row_number in range(2, last_row + 1):
        sheet.row_dimensions[row_number].height = 18
    for column, header in enumerate(headers, 1):
        letter = get_column_letter(column)
        sample = [header] + [row[column - 1] for row in rows[:500]]
        sheet.column_dimensions[letter].width = min(48, max(10, max(len(value) for value in sample) + 2))
    if last_column < MAX_EXCEL_COLUMNS:
        sheet.column_dimensions.group(get_column_letter(last_column + 1), "XFD", hidden=True)

    output = io.BytesIO()
    workbook.save(output)
    safe_stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", Path(filename).stem).strip(" .") or "client-data"
    return output.getvalue(), f"{safe_stem}-client-shareable.xlsx"
